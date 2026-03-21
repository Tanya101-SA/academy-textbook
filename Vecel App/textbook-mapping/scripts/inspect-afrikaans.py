"""Inspect Textbook Topic Mapping v2.xlsx for Afrikaans vs English rows per grade."""
import openpyxl

path = 'C:/Users/NicolaCrous/OneDrive - Speccon Holdings (Pty) Ltd/SpecCon Academy - Textbooks/Textbook Topic Mapping v2.xlsx'
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

AF_WORDS = [
    'rekeningkunde', 'boekhouding', 'finansi', 'etiek', 'beginsels',
    'inkomstestaat', 'balansstaat', 'waardevermindering', 'joernale', 'grootboek',
    'debiteure', 'krediteure', 'vennootskap', 'maatskappy', 'wiskunde',
    'geskiedenis', 'aardrykskunde', 'lewenswetenskappe', 'meetkunde',
    'taalstrukture', 'instruksies', 'aanwysings', 'gedig', 'nuusberig',
    'luisterbegrip', 'leesbegrip', 'hardoplees', 'mondelinge', 'hervertelling',
    'dagboekinskrywing', 'werkwoorde', 'naamwoorde', 'voornaamwoorde',
    'lidwoorde', 'meervoude', 'sinonieme', 'antonieme', 'leestekens',
    'skryftekens', 'punktuasie', 'byvoeglike', 'voorsetsels', 'voegwoorde',
    'beskrywende', 'verhalende', 'opstel', 'rolspel', 'bespreking',
    'tydsvorme', 'ontkenning', 'lydende', 'bedrywende', 'direkte', 'indirekte',
    'sinsoorte', 'sinstrukture', 'sinsuitbreiding', 'woordsoorte', 'selfstandige',
    'bywoorde', 'lees en kyk', 'luister en praat', 'skryf en aanbied',
    'alfabetiese', 'vokale', 'konsonante', 'afkortings', 'trappe van vergelyking',
    'intensiewe', 'idiome', 'lettergrepe', 'klankgrepe', 'afleidings',
    'samestellings', 'spreekwoorde', 'eufemisme', 'figuurlike',
    'alliterasie', 'assonansie', 'personifikasie',
    'kontantvloei', 'ontleding', 'aandele', 'verdeling', 'voorraad',
    'belasting', 'skuld', 'kapitaal', 'bates', 'laste',
    'definisie', 'verduideliking', 'prosesse', 'eienskappe', 'verskynsels',
]

total_af = 0
total_en = 0

for sn in wb.sheetnames:
    if sn == 'Summary':
        continue
    ws = wb[sn]
    section = ''
    af = 0
    en = 0
    sections_af = {}
    sections_en = {}
    for row in ws.iter_rows(values_only=True):
        c0 = str(row[0]).strip() if row[0] else ''
        c1 = str(row[1]).strip() if row[1] else ''
        c2 = str(row[2]) if row[2] else ''
        if c0.startswith('Grade') and not c1:
            section = c0
        elif c0.startswith('Term') and c1 and c1 != 'LAP Topic':
            text = c1 + ' ' + c2
            is_af = ("'n " in text or '\u2019n ' in text or
                     any(w in text.lower() for w in AF_WORDS))
            if is_af:
                af += 1
                sections_af[section] = sections_af.get(section, 0) + 1
            else:
                en += 1
                sections_en[section] = sections_en.get(section, 0) + 1
    total_af += af
    total_en += en
    print(f'\n=== {sn}: {af} AF rows, {en} EN rows ===')
    for sec in sorted(set(list(sections_af.keys()) + list(sections_en.keys()))):
        a = sections_af.get(sec, 0)
        e = sections_en.get(sec, 0)
        tag = ' <-- AF' if a > 0 else ''
        print(f'  {sec}: {a} AF, {e} EN{tag}')

print(f'\n=== TOTALS: {total_af} AF rows, {total_en} EN rows ===')
